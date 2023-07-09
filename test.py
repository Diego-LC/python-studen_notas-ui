""" import random
import string

code_str = string.ascii_letters + string.digits
## Imprime 4 letras o números aleatorios
print(''.join(random.sample(code_str,6)))
p = 'asdfsdsd@gmail.com'
print('@' in p or '.c' in p) """

#ventana estudiante usando POO
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import utils

def crear_exel(datos):
    libro_excel = Workbook()
    hoja = libro_excel.active

    # Encabezados de las notas
    encabezados_notas = ['Tipo Nota', 'Nota', 'Ponderación Nota','Tipo Evaluación']
    hoja.append(encabezados_notas)

    # Escribir notas
    for fila_data in datos['notas'].values():
        fila = [fila_data['Tipo Nota'], fila_data['Nota'], fila_data['Ponderación Nota'],fila_data['Tipo Evaluación']]
        hoja.append(fila)

    # Espacio en blanco entre las notas y las ponderaciones
    hoja.append([])

    # Encabezados de las ponderaciones
    encabezados_ponderaciones = ['Tipo Evaluación', 'Ponderación']
    hoja.append(encabezados_ponderaciones)

    # Escribir ponderaciones
    for tipo_evaluacion, ponderacion in datos['ponderaciones']['tipo_evaluacion'].items():
        fila = [tipo_evaluacion, ponderacion]
        hoja.append(fila)

    # Espacio en blanco entre las ponderaciones y los promedios
    hoja.append([])
    hoja.append(['Promedios'])

    # Encabezados y valores de los promedios por tipo de evaluación
    encabezados_promedios = ['Tipo Evaluación', 'Promedio']
    hoja.append(encabezados_promedios)

    for tipo_evaluacion, promedio in datos['promedios'].items():
        fila = [tipo_evaluacion, promedio]
        hoja.append(fila)
    hoja.append([])

    # Encabezado y valor del promedio total
    encabezado_promedio_total = ['Promedio Total', datos['promedio_total']]
    hoja.append(encabezado_promedio_total)

    # Ajustar el ancho de las columnas
    for columna in hoja.columns:
        max_length = 0
        column = columna[0].column_letter  # Obtiene la letra de la columna (A, B, C, ...)
        for cell in columna:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        hoja.column_dimensions[column].width = adjusted_width

    # Centrar el contenido de las celdas
    for fila in hoja.iter_rows(min_row=2):
        for celda in fila:
            celda.alignment = Alignment(horizontal="center", vertical="center")
    
    ruta = "notas.xlsx"
    libro_excel.save(ruta)
    return ruta

datos = {
    'userid': '64a9f9ea2c54975231b56c04',
    'notas': {
        'fila1': {'Tipo Nota': 'Prueba', 'Nota': '4.8', 'Ponderación Nota': '0.4', 'Tipo Evaluación': 'Eval. Teórica'}, 
        'fila2': {'Tipo Nota': 'Control', 'Nota': '6.5', 'Ponderación Nota': '0.2', 'Tipo Evaluación': 'Eval. Teórica'}, 
        'fila3': {'Tipo Nota': 'Prueba', 'Nota': '5.7', 'Ponderación Nota': '0.4', 'Tipo Evaluación': 'Eval. Teórica'}, 
        'fila4': {'Tipo Nota': 'Trabajo', 'Nota': '6.3', 'Ponderación Nota': '1.0', 'Tipo Evaluación': 'Eval. Práctica'}
        },
    'ponderaciones': {
        'tipo_evaluacion': {'Eval. Práctica': '0.4', 'Eval. Teórica': '0.6'},
        'tipo_nota': {'Control': '0.2', 'Prueba': '0.4', 'Trabajo': '1.0'}
        },
    'promedios': {'Eval. Teórica': 5.5, 'Eval. Práctica': 6.3},
    'promedio_total': 5.82
}

ruta = crear_exel(datos)
print(ruta)


""" # Escribir los datos en el archivo Excel
    hoja.append(['Tipo Nota', 'Nota', 'Tipo Evaluación'])

    for fila in datos['notas'][1]['notas']:
        hoja.append(fila)

    hoja.append([])  # Agregar una fila vacía
    hoja.append(['Ponderaciones:'])

    # Obtener la columna inicial para las ponderaciones
    columna_inicial = utils.get_column_letter(3)  # Columna 'C'

    for clave, valor in datos['notas'][0]['ponderaciones'].items():
        hoja.append([clave])
        for subclave, subvalor in valor.items():
            hoja.append([subclave, subvalor])

    hoja.append([])  # Agregar una fila vacía

    # Obtener la columna inicial para los promedios
    columna_promedios = 3  # Columna siguiente a la última columna utilizada
    fila_promedios = 7  # Fila siguiente a la última fila utilizada
    hoja.cell(row=6, column=columna_promedios).value = 'Promedios:'

    for clave, valor in datos['promedios'].items():
        hoja.cell(row=fila_promedios, column=columna_promedios).value = clave
        hoja.cell(row=fila_promedios, column=columna_promedios + 1).value = valor
        fila_promedios += 1
 """
